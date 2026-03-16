import math
import os
import shutil
import matplotlib.pyplot as plt
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from matplotlib.patches import Patch
from io import BytesIO
from matplotlib.lines import Line2D
import matplotlib.patches as patches
from io import BytesIO
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.datavalidation import DataValidation

COLORS = ["EFF8E6", "FBFBEF", "F3F3F3", "DFEDF4", "E8E5F3", "FDEEF5"]
LOUVER_PROFILE_COLUMNS = {
    "Area Name": 1,
    "S. No": 2,
    "Length (mm)": 3,
    "Piece Qty (nos)": 4,
    "Total Product Length (m)": 5,
    "Orientation": 6,
    "Height (mm)": 7,
    "Width (mm)": 8,
    "Area Qty (nos)": 9,
    "Area (ft2)": 10,
    "No. of Pieces": 11,
    "Total Length (m)": 12,
    "Scrap Length (mm)": 13,
    "Scrap Qty (nos)": 14,
}
BEAMC_PROFILE_COLUMNS = {
    "Area Name": 1,
    "S. No": 2,
    "Beam Length (mm)": 3,
    "Piece Qty (nos)": 4,
    "Total Product Length (m)": 5,
    "Width (mm)": 6,
    "Length (mm)": 7,
    "Area Qty (nos)": 8,
    "Total Length (m)": 9,
    "Scrap Length (mm)": 10,
    "Scrap Qty (nos)": 11,
}
CARRIERS_BY_PRODUCT = {
    "Grille 2550": [
        ("Total Carriers Length (m)", "Carrier"),
        ("Total Covering Length (m)", "Covering Plate"),
    ],
    "S-Louvers": [("133.7 mm Carrier Pieces (nos)", "Carrier Pieces")],
    "Rectangular Louvers": [("Total Carrier Length (m)", "Carrier")],
    "Fluted": [("Aluminum Pipe Length (m)", "Aluminum Pipe")],
    "Cottal": [("Aluminum Pipe Length (m)", "Aluminum Pipe")],
    "Beam C-Channel": [],
    "Aerofoil": [],
}
TITLE_ROWS = {"profile": 4, "carrier": 2, "offer": 3}
NON_NUMERIC_COLS = [
    "Used Table",
    "Waste Table",
    "Orientation",
    "L End Cap Orientation",
    "Area Name",
    "S. No",
    "Project Title",
    "Aluminum Pipe Grade",
    "Aluminum Pipe Length (m)",
    "Aerofoil Type",
    "MS Rod Lengths (mm)",
    "Total MS Rods for each Length (pcs)",
]

# Updated INV_COLUMNS with Product Code restored
INV_COLUMNS = [
    "Product Code",
    "Product Name",
    "Length",
    "Quantity",
    "UOM",
    "Colour",
    "Finish",
    "CNC Hole Distance",
    "Remarks",
]

# Updated PIPE_MAPPER with codes restored
PIPE_MAPPER = {
    "50x25": ("AC-RT-5025-02", "RECTANGULAR TUBE 50x25"),
    "38x25": ("RT-3825-01", "RECTANGULAR TUBE 38x25"),
    "25x12": ("RT-2512-01", "RECTANGULAR TUBE 25x12"),
}

# Common accessory codes
COMMON_ACCESSORIES = {
    "BLACK_GYPSUM_19MM": ("AC-SC-DW19", "BLACK GYPSUM DRYWALL SCREW 19MM"),
    "FULL_THREADED_75MM": ("AC-SC-FT75", "FULL THREADED SCREW 75MM"),
    "PVC_GITTY_50X10MM": ("AC-GT-1050", "PVC GITTY 50X10MM"),
    "SELF_DRILLING_19MM": ("AC-SC-SD19", "SELF DRILLING SCREW 19MM"),
    "SELF_DRILLING_25MM": ("AC-SC-SD25", "SELF DRILLING SCREW 25MM"),
    "RIVET_6MM": ("AC-RV-0006", "RIVET 6MM"),
    "PAINT": ("AC-GN-PT", "PAINT"),
    "PAINT_BRUSH": ("AC-GN-PB", "PAINT BRUSH"),
    "EPDM_GASKET": ("AC-GN-ER", "EPDM RUBBER GASKET"),
    "NUT_3_5X25": ("AC-NT-3525", "NUT 3.5x25"),
    "BOLT_3_5X25": ("AC-BT-3525", "BOLT 3.5x25"),
    "WASHER": ("AC-WS-3525", "PATTA WASHER"),
    "TRUSS_HEAD_16MM": ("AC-SC-TP16", "TRUSS HEAD PHILLIPS SCREW 16MM"),
    "L_ANGLE_19X19": ("AC-LA-1919", "L-ANGLE 19x19"),
    "L_ANGLE_25X25": ("AC-LA-2525", "L-ANGLE 25x25"),
    "GRILLE_CARRIER": ("GR-CA-01", "GRILLE CARRIER"),
}

PVC_GITTY_TYPES = [
    ("AC-GT-1030", "PVC GITTY 30X10MM"),
    ("AC-GT-1040", "PVC GITTY 40X10MM"),
    ("AC-GT-1050", "PVC GITTY 50X10MM"),
]

FT_TYPES = [
    ("AC-SC-FT19", "FULL THREADED SCREW 19MM"),
    ("AC-SC-FT25", "FULL THREADED SCREW 25MM"),
    ("AC-SC-FT38", "FULL THREADED SCREW 38MM"),
    ("AC-SC-FT50", "FULL THREADED SCREW 50MM"),
    ("AC-SC-FT63", "FULL THREADED SCREW 63MM"),
    ("AC-SC-FT75", "FULL THREADED SCREW 75MM"),
]

SELF_DRILLING_TYPES = [
    ("AC-SC-SD19", "SELF DRILLING SCREW 19MM"),
    ("AC-SC-SD25", "SELF DRILLING SCREW 25MM"),
    ("AC-SC-SD38", "SELF DRILLING SCREW 38MM"),
    ("AC-SC-SD50", "SELF DRILLING SCREW 50MM"),
    ("AC-AC-RV6", "RIVET 6MM"),
]

COTTAL_CORNERS = [
    ("CT-C1-02", "COTTAL CORNER 1"),
    ("CT-C2-02", "COTTAL CORNER 2"),
    ("CT-C3-02", "COTTAL CORNER 3"),
    ("CT-C4-02", "COTTAL CORNER 4"),
]

L_ANGLE = [
    ("AC-LA-1919", "L-ANGLE 19x19"),
    ("AC-LA-2525", "L-ANGLE 25x25"),
    ("AC-LA-2550", "L-ANGLE 25x50"),
    ("AC-LA-5050", "L-ANGLE 50x50"),
]

BLACK_GYPSUM = [
    ("AC-SC-DW12", "BLACK GYPSUM DRYWALL SCREW 12MM"),
    ("AC-SC-DW19", "BLACK GYPSUM DRYWALL SCREW 19MM"),
    ("AC-SC-DW25", "BLACK GYPSUM DRYWALL SCREW 25MM"),
    ("AC-SC-DW38", "BLACK GYPSUM DRYWALL SCREW 38MM"),
]
FIXING_TYPES = BLACK_GYPSUM + [("AC-RV-0006", "RIVET 6MM")]
RECT_EC_FIXING = [
    ("AC-SC-DW19", "BLACK GYPSUM DRYWALL SCREW 19MM"),
    ("AC-SC-SSMC", "SS MACHINE SCREW"),
]


def get_xl_templates(product, dir):
    path = f"reference_xls/{dir}_templates"
    XL_TEMPLATES = {
        "Grille 2550": f"{path}/grille.xlsx",
        "Cottal": f"{path}/cottal.xlsx",
        "Fluted": f"{path}/fluted.xlsx",
        "Aerofoil-C-Channel": f"{path}/aerofoil_c_channel.xlsx",
        "Aerofoil-Fringe End Caps": f"{path}/aerofoil_endcap.xlsx",
        "Aerofoil-MS Rod/Slot Cut Pipe": f"{path}/aerofoil_slot_pipe.xlsx",
        "Aerofoil-D-Wall Bracket": f"{path}/aerofoil_d_bracket.xlsx",
        "Aerofoil-Moveable (Manual)": f"{path}/aerofoil_manual_moveable.xlsx",
        "S-Louvers": f"{path}/s_louver.xlsx",
        "Rectangular Louvers": f"{path}/rectangular.xlsx",
        "Beam C-Channel": f"{path}/beam_c_channel.xlsx",
    }
    return XL_TEMPLATES[product]


def value(series, column_name, qty_windows=1):
    if column_name in NON_NUMERIC_COLS:
        return series[column_name]
    elif column_name in ["Height (mm)", "Width (mm)", "Area Qty (nos)", "Pitch (mm)"]:
        return int(series[column_name])
    return float(series[column_name]) * qty_windows


def merge_cells(xl, row_start, row_end, col_start, col_end):
    xl.merge_cells(
        start_row=row_start, start_column=col_start, end_row=row_end, end_column=col_end
    )


def get_col_ref(col):
    return f"{chr(64 + col)}"


def set_cell(sheet_type, cell, value, bold=False, alignment="center", size=10):

    font_map = {
        "offer": "Montserrat",
        "profile": "Times New Roman",
        "inventory": "Times New Roman",
        "installer": "Helvetica"
    }
    row_height_map = {
        "offer": 30
    }

    font = font_map[sheet_type]

    if isinstance(value, float):
        value = round(value, 2)

    cell.value = value
    font = Font(name=font, size=size, bold=bold)
    cell.font = font
    cell.alignment = Alignment(
        horizontal=alignment,
        vertical="center",
        wrap_text=True
    )

    if sheet_type in row_height_map:
        cell.parent.row_dimensions[cell.row].height = row_height_map[sheet_type]


def set_sum_formula(cell, xl_type):
    col_ref = get_col_ref(cell.column)
    formula = '=SUM(${}${}:INDIRECT("{}" & ROW()-1))'.format(
        col_ref, TITLE_ROWS[xl_type] + 1, col_ref
    )
    if xl_type == "offer":
        set_cell("offer", cell, formula, bold=True)
        cell.parent.row_dimensions[cell.row].height = 30
    else:
        set_cell(cell, formula, bold=True)


def add_box_border(ws, style, start_row, end_row, start_col, end_col):
    thick_side = Side(style=style)

    for col in range(start_col, end_col + 1):
        # Top border
        cell = ws.cell(row=start_row, column=col)
        cell.border = Border(
            top=thick_side,
            left=cell.border.left,
            right=cell.border.right,
            bottom=cell.border.bottom,
        )

        # Bottom border
        cell = ws.cell(row=end_row, column=col)
        cell.border = Border(
            bottom=thick_side,
            left=cell.border.left,
            right=cell.border.right,
            top=cell.border.top,
        )

    for row in range(start_row, end_row + 1):
        # Left border
        cell = ws.cell(row=row, column=start_col)
        cell.border = Border(
            left=thick_side,
            top=cell.border.top,
            right=cell.border.right,
            bottom=cell.border.bottom,
        )

        # Right border
        cell = ws.cell(row=row, column=end_col)
        cell.border = Border(
            right=thick_side,
            top=cell.border.top,
            left=cell.border.left,
            bottom=cell.border.bottom,
        )

    return ws


def adjust_cell(cell, threshold=35):

    text_length = len(str(cell.value))
    lines = 1
    if text_length > threshold:
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        lines = math.ceil(text_length / threshold)
        new_height = 19 * (lines + 1)
        cell.parent.row_dimensions[cell.row].height = new_height


def add_total_border(xl, row, col_start, col_end):

    top_border = Border(
        top=Side(style="thin", color="FD5300")
    )

    for col in range(col_start, col_end + 1):
        cell = xl.cell(row=row, column=col)

        cell.border = Border(
            top=top_border.top
        )

        adjust_cell(cell)

    return xl


def add_total_cell(xl, row, col_start, col_end, xl_type="general"):
    merge_cells(xl, row, row, col_start, col_end)
    total_cell = xl.cell(row=row, column=1)

    if xl_type == "offer":
        set_cell("offer", total_cell, "Total", bold=True, alignment="center")
    else:
        set_cell(total_cell, "Total", bold=True, alignment="left")
    return xl


def add_total_row(xl, xl_type, row_max, col_max):

    product_cell = xl.cell(row=1, column=1)
    product = product_cell.value

    if xl_type == "profile":
        for idx in range(1, col_max + 1):
            col_title = xl.cell(row=TITLE_ROWS[xl_type], column=idx).value
            if col_title in [
                "Area Qty (nos)",
                "Qty (nos)",
                "Total Length (m)",
                "Area (ft2)",
                "No. of Pieces",
            ]:
                cell = xl.cell(row=row_max, column=idx)
                set_sum_formula(cell, xl_type)

        # xl = add_box_border(xl, "medium", row_max, row_max, 1, col_max)
        xl = add_total_cell(xl, row_max, 1, 3)
        merge_cells(xl, row_max, row_max, 6, 8)

        return xl
    else:
        for idx in range(col_max, 0, -1):
            col_title = xl.cell(row=TITLE_ROWS[xl_type], column=idx).value

            if product == "Beam C-Channel":
                if col_title == "Area Qty (nos)":
                    break
            else:
                if col_title == "Width (mm)":
                    break

            cell = xl.cell(row=row_max, column=idx)
            set_sum_formula(cell, xl_type)

        if xl_type == "offer":
            xl = add_total_border(xl, row_max, 1, col_max)
            xl = add_total_cell(xl, row_max, 1, idx, xl_type="offer")
        else:
            xl = add_box_border(xl, "thin", row_max, row_max, 1, col_max)
            xl = add_total_cell(xl, row_max, 1, idx)

        return xl


def remove_zero_total_cols(xl, zero_cols, col_max, title_only, offer_df_cols):

    titles = {}

    # Unmerge titles
    if title_only:
        for i in range(1, 3):
            xl.unmerge_cells(start_row=i, end_row=i, start_column=3, end_column=col_max)
            titles[i] = xl.cell(row=i, column=3).value
    else:
        xl.unmerge_cells(start_row=1, end_row=2, start_column=9, end_column=col_max)
        titles[1] = xl.cell(row=1, column=9).value

    col_names = list(offer_df_cols.keys())

    for col in range(col_max, 0, -1):
        df_col = col_names[col - 1]

        if df_col in zero_cols:
            xl.delete_cols(col)
            col_max -= 1

    # Re-merge titles
    if title_only:
        for i in range(1, 3):
            xl.merge_cells(start_row=i, end_row=i, start_column=3, end_column=col_max)
            cell = xl.cell(row=i, column=3)
            set_cell("offer", cell, titles[i], bold=True, size=12)
    else:
        xl.merge_cells(start_row=1, end_row=2, start_column=9, end_column=col_max)
        cell = xl.cell(row=1, column=9)
        set_cell("offer", cell, titles[1], bold=True, size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    return xl, col_max


# def fill_window_data(product, start, title_row, xl, results, xl_type="general"):

#     non_qty_cols = {
#         "Area Name",
#         "S. No",
#         "Orientation",
#         "Height (mm)",
#         "Width (mm)",
#         "Length (mm)",
#         "Area Qty (nos)",
#         "Aluminum Pipe Grade",
#         "C-Plate Width",
#         "Length per L-Angle (mm)",
#         "MS Rod Lengths (mm)",
#         "Total MS Rods for each Length (pcs)",
#         "Start Piece (29-3002-00) Piece Length (mm)",
#         "End Pieces (29-4001-00) Piece Length (mm)",
#         "End Piece Length (mm)",
#         "L End Cap Orientation",
#     }
#     curr_row = start

#     window_qty_col = 6
#     if product == "Beam C-Channel":
#         window_qty_col = 5

#     for i in range(len(results)):
#         window_data = results.iloc[i]

#         for col in range(1, 25):
#             col_title = xl.cell(row=title_row, column=col).value
#             if col_title is None:
#                 first_none_col = col
#                 break

#             cell = xl.cell(row=curr_row, column=col)
#             if cell.data_type == "f":
#                 continue

#             val = value(window_data, col_title)

#             if val == 0:
#                 cell.value = "x"
#                 font = Font(size=14, bold=True, color="FF0000")
#                 cell.font = font
#                 cell.alignment = Alignment(horizontal="center", vertical="center")
#             else:
#                 if xl_type == "offer":
#                     if col_title in non_qty_cols:
#                         set_cell("offer", cell, val)
#                     else:
#                         set_cell("offer", cell, f"={val}*{get_col_ref(window_qty_col)}{curr_row}")
#                 else:
#                     if col_title in non_qty_cols:
#                         set_cell(cell, val)
#                     else:
#                         set_cell(cell, f"={val}*{get_col_ref(window_qty_col)}{curr_row}")
#         curr_row += 1

#     return xl, window_data, curr_row, first_none_col - 1


def fill_offer_data(xl, start_row, offer_df, offer_df_cols):

    curr_row = start_row

    # Find position of qty_areas in your ordered dict
    col_names = list(offer_df_cols.keys())
    qty_col_index = col_names.index("qty_areas") + 1

    for _, row in offer_df.iterrows():
        col_idx = 1

        for col_name, col_info in offer_df_cols.items():
            col_type = col_info["type"]
            cell = xl.cell(row=curr_row, column=col_idx)
            val = row[col_name]

            if col_type == "desc":
                set_cell("offer", cell, val)

            elif col_type == "formula":
                qty_ref = f"{get_col_ref(qty_col_index)}{curr_row}"
                formula = f"={val}*{qty_ref}"
                set_cell("offer", cell, formula)

            col_idx += 1

        curr_row += 1

    return xl, curr_row, len(col_names)


def add_dropdown(xl, lst):
    dv = DataValidation(
        type="list", formula1='"' + ",".join(lst) + '"', allow_blank=True
    )
    xl.add_data_validation(dv)
    return xl, dv


def generate_window_image(row, common_vars):

    plt.rcdefaults()

    width  = row["width"]
    height = row["height"]
    orient = row["orientation"]
    cuts   = row["cut_summary"]
    carrier_distances = row.get("carrier_distances", [])

    N_BARS = 20

    fig, ax = plt.subplots(figsize=(10, 8))
    if height >= width:
        W, H = 300, 380
    else:
        W, H = 380, 300

    ax.set_xlim(-120, W + 220)
    ax.set_ylim(-70, H + 70)
    ax.set_aspect("equal")
    ax.axis("off")

    # ── Window frame ──
    frame = patches.Rectangle(
        (0, 0), W, H,
        linewidth=2, edgecolor="#1a1a1a", facecolor="white", zorder=2
    )
    ax.add_patch(frame)

    # ── Right side info panel ──
    pitch        = common_vars.get("pitch", "")
    divisions    = row.get("divisions", "")

    cut_summary = row.get("cut_summary", "")
    if len(cut_summary) > 1:
        cut_summary_str = "+".join(str(c) for c in cut_summary)
    else:
        cut_summary_str = f"Single Piece {cut_summary[0]} mm"

    direction = row["louver_direction"]

    info_lines = [
        (f"{divisions} Divisions",    "#333",   12,  True),
        (f"@ {pitch} mm Pitch",       "#333",   12,  True),
        ("",       "#333",   12,  True),
        ("Breakdown",        "#333",   12,  False),
        (f"{cut_summary_str}",        "#333",   12,  True),
        ("Grille Direction",        "#333",   12,  False),
        (f"{direction}",        "#333",   12,  True),
    ]

    line_gap = 22
    info_x = W + 180
    start_y = H * 0.7
    for i, (text, color, size, bold) in enumerate(info_lines):
        ax.text(
            info_x, start_y - i * line_gap,
            text,
            fontsize=size,
            ha="center", va="top",
            color=color,
            fontweight="bold" if bold else "normal",
        )

    # Build cumulative joint positions
    total = sum(cuts)
    cumulatives = []
    c = 0
    for cut in cuts:
        c += cut
        cumulatives.append(c)

    ENDCAP_SIZE = 6
    # Map direction to which sides are active
    endcap_choice = str(row.get("endcaps", "")).strip()
    louver        = str(row.get("louver_direction", "")).strip()
    num_L = row['num_L']
    num_inverse_L = row['num_inverse_L']

    if orient == "Vertical":
        if louver == "Top L":
            draw_top    = endcap_choice in ("Both sides", "Single Side - L")         and num_L > 0
            draw_bottom = endcap_choice in ("Both sides", "Single Side - Inverse L") and num_inverse_L > 0
        elif louver == "Top Inverse L":
            draw_top    = endcap_choice in ("Both sides", "Single Side - Inverse L") and num_inverse_L > 0
            draw_bottom = endcap_choice in ("Both sides", "Single Side - L")         and num_L > 0
        else:
            draw_top    = False
            draw_bottom = False
        draw_left  = False
        draw_right = False

    else:  # Horizontal
        if louver == "Right L":
            draw_right = endcap_choice in ("Both sides", "Single Side - L")         and num_L > 0
            draw_left  = endcap_choice in ("Both sides", "Single Side - Inverse L") and num_inverse_L > 0
        else:
            draw_left  = False
            draw_right = False
        draw_top    = False
        draw_bottom = False

    def endcap(x, y):
        ax.add_patch(patches.Rectangle(
            (x - ENDCAP_SIZE/2, y - ENDCAP_SIZE/2),
            ENDCAP_SIZE, ENDCAP_SIZE,
            linewidth=0.5, edgecolor="#c4b800",
            facecolor="#F7E301", zorder=5
        ))

    if orient == "Vertical":
        # ── Bars ──
        for i in range(N_BARS):
            x = (i + 0.5) * W / N_BARS
            if draw_top:    endcap(x, H)
            if draw_bottom: endcap(x, 0)
            ax.plot([x, x], [0, H], color="#888", linewidth=1, zorder=3)

        # ── Joint lines + cumulative labels (right side) ──
        for cum in cumulatives:
            y_pos = (cum / total) * H
            is_edge = (cum == total)
            if not is_edge:
                ax.plot([0, W], [y_pos, y_pos],
                        color="#e63946", linewidth=1.8, zorder=4)
            ax.text(W + 8, y_pos, f"{cum}", fontsize=12,
                    va="center", ha="left", color="#e63946", fontweight="bold")

        # 0 at bottom right
        ax.text(W + 8, 0, "0", fontsize=12,
                va="center", ha="left", color="#e63946", fontweight="bold")

        # ── Piece length labels (right side, between joints) ──
        prev_y = 0
        for cut, cum in zip(cuts, cumulatives):
            y_pos = (cum / total) * H
            mid_y = (prev_y + y_pos) / 2
            ax.text(W + 8, mid_y, f"← {cut}", fontsize=12,
                    va="center", ha="left", color="#555", fontweight="bold")
            prev_y = y_pos

        # ── Carrier lines + labels (left side) ──
        piece_start_y = 0
        for piece_idx, (cut, cum) in enumerate(zip(cuts, cumulatives)):
            piece_end_y = (cum / total) * H
            piece_h_px  = piece_end_y - piece_start_y
            piece_len   = cut

            ax.text(-8, piece_start_y, "0", fontsize=12,
                    va="center", ha="right", color="#2FB6DE", fontweight="bold")

            if piece_idx < len(carrier_distances):
                for carrier_pos in carrier_distances[piece_idx]:
                    y_carrier = piece_start_y + (carrier_pos / piece_len) * piece_h_px
                    ax.plot([0, W], [y_carrier, y_carrier],
                            color="#2FB6DE", linewidth=1.4,
                            linestyle="--", zorder=4, alpha=0.85)
                    ax.text(-8, y_carrier, f"{carrier_pos}", fontsize=12,
                            va="center", ha="right", color="#2FB6DE", fontweight="bold")

            piece_start_y = piece_end_y

    else:  # Horizontal
        # ── Bars ──
        for i in range(N_BARS):
            y = (i + 0.5) * H / N_BARS
            if draw_left:  endcap(0, y)
            if draw_right: endcap(W, y)
            ax.plot([0, W], [y, y], color="#888", linewidth=1, zorder=3)

        # ── Joint lines + cumulative labels (bottom) ──
        for cum in cumulatives:
            x_pos = (cum / total) * W
            is_edge = (cum == total)
            if not is_edge:
                ax.plot([x_pos, x_pos], [0, H],
                        color="#e63946", linewidth=1.8, zorder=4)
            ax.text(x_pos, -10, f"{cum}", fontsize=12,
                    ha="center", va="top", color="#e63946", fontweight="bold")

        # 0 at bottom left
        ax.text(0, -10, "0", fontsize=12,
                ha="center", va="top", color="#e63946", fontweight="bold")

        # ── Piece length labels (below joint labels) ──
        prev_x = 0
        for cut, cum in zip(cuts, cumulatives):
            x_pos = (cum / total) * W
            mid_x = (prev_x + x_pos) / 2
            ax.text(mid_x, -22, f"{cut}", fontsize=12,
                    ha="center", va="top", color="#555", fontweight="bold")
            prev_x = x_pos

        # ── Carrier lines + labels (top side) ──
        piece_start_x = 0
        for piece_idx, (cut, cum) in enumerate(zip(cuts, cumulatives)):
            piece_end_x = (cum / total) * W
            piece_w_px  = piece_end_x - piece_start_x
            piece_len   = cut

            ax.text(piece_start_x, H + 8, "0", fontsize=10,
                    va="bottom", ha="center", color="#2FB6DE",
                    fontweight="bold", rotation=45)

            if piece_idx < len(carrier_distances):
                for carrier_pos in carrier_distances[piece_idx]:
                    x_carrier = piece_start_x + (carrier_pos / piece_len) * piece_w_px
                    ax.plot([x_carrier, x_carrier], [0, H],
                            color="#2FB6DE", linewidth=1.4,
                            linestyle="--", zorder=4, alpha=0.85)
                    ax.text(x_carrier, H + 8, f"{carrier_pos}", fontsize=10,
                            va="bottom", ha="center", color="#2FB6DE",
                            fontweight="bold", rotation=45)

            piece_start_x = piece_end_x

    # ── Width annotation (top) ──
    ax.annotate("", xy=(W, H + 45), xytext=(0, H + 45),
            arrowprops=dict(arrowstyle="<->", color="#333", lw=1.2))
    ax.text(W / 2, H + 50, f"Width: {width} mm",
            ha="center", va="bottom", fontsize=10,
            color="#333", fontweight="bold")

    # ── Height annotation (left) ──
    ax.annotate("", xy=(-80, H), xytext=(-80, 0),
                arrowprops=dict(arrowstyle="<->", color="#333", lw=1.2))
    ax.text(-90, H / 2, f"Height: {height} mm",
            ha="right", va="center", fontsize=12,
            color="#333", fontweight="bold", rotation=90)

    legend_elements = [
        Line2D([0], [0], color="#888",    linewidth=1.5, label=common_vars.get("product", "Profile")),
        Line2D([0], [0], color="#e63946", linewidth=1.8, label="Joint Line"),
        Line2D([0], [0], color="#2FB6DE", linewidth=1.4, linestyle="--", label="Carrier"),
        Patch(facecolor="#F7E301", edgecolor="#c4b800", label="Endcap"),
    ]

    ax.legend(
        handles=legend_elements,
        loc="upper right",
        fontsize=12,
        frameon=True,
        framealpha=0.9,
        edgecolor="#ccc",
        facecolor="white",
        markerscale=2,
        handlelength=2.5,
        handleheight=1.5,
    )

    plt.tight_layout(pad=0.5)
    buf = BytesIO()
    plt.savefig(buf, format="png", dpi=150, facecolor="white")
    plt.close(fig)
    buf.seek(0)
    return buf


def get_libreoffice_path():
    # Check PATH first
    lo = shutil.which("libreoffice") or shutil.which("soffice")
    if lo:
        return lo
    # macOS default install locations
    for path in [
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
        "/Applications/LibreOffice.app/Contents/MacOS/libreoffice",
    ]:
        if os.path.exists(path):
            return path
    raise FileNotFoundError(
        "LibreOffice not found. Install via: brew install --cask libreoffice"
    )


def fill_cut_plan(ws, cut_plan, start_row=70):
    from openpyxl.styles import Border, Side

    thin   = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def apply_border(cell):
        cell.border = border

    current_row = start_row

    for entry in cut_plan:
        stock_length = entry["stock_length"]
        stock_qty    = entry["qty"]
        cuts         = entry["cuts"]  # e.g. [1375, 1375, 1375]

        n = len(cuts)  # one row per cut piece

        # Merge stock length and stock qty across all cut rows
        if n > 1:
            ws.merge_cells(f"C{current_row}:C{current_row + n - 1}")
            ws.merge_cells(f"D{current_row}:D{current_row + n - 1}")

        set_cell("installer", ws[f"C{current_row}"], stock_length, alignment="center")
        set_cell("installer", ws[f"D{current_row}"], stock_qty,    alignment="center")

        for i, cut_len in enumerate(cuts):
            row = current_row + i
            set_cell("installer", ws[f"E{row}"], cut_len,   alignment="center")
            set_cell("installer", ws[f"F{row}"], stock_qty, alignment="center")

            for col in ["C", "D", "E", "F"]:
                apply_border(ws[f"{col}{row}"])

        current_row += n
