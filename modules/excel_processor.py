from datetime import date
from io import BytesIO

import openpyxl
import pandas as pd
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.drawing.image import Image as XLImage

import copy
import subprocess
import os
from openpyxl import load_workbook
from pypdf import PdfWriter, PdfReader

import streamlit as st

from modules.excel_utils import (
    BEAMC_PROFILE_COLUMNS,
    BLACK_GYPSUM,
    COLORS,
    COTTAL_CORNERS,
    FIXING_TYPES,
    FT_TYPES,
    INV_COLUMNS,
    L_ANGLE,
    LOUVER_PROFILE_COLUMNS,
    PVC_GITTY_TYPES,
    RECT_EC_FIXING,
    SELF_DRILLING_TYPES,
    add_box_border,
    add_dropdown,
    add_total_row,
    fill_cut_plan,
    generate_window_image,
    get_libreoffice_path,
    # fill_window_data,
    get_xl_templates,
    remove_zero_total_cols,
    set_cell,
    value,
    fill_offer_data
)


# def generate_profile_xl(start, profile_xl, results):

#     product = "Louvers"
#     if profile_xl.cell(row=1, column=1).value == "Beam C-Channel":
#         product = "Beam C-Channel"

#     # Sort results by serial number for better readability
#     results["S. No"].replace("", pd.NA, inplace=True)
#     results.sort_values(by="S. No", ascending=True, na_position="last", inplace=True)
#     results["S. No"].replace(pd.NA, "", inplace=True)

#     # Assign every serial group a unique color
#     results["S. No Group"] = results["S. No"].str.strip().str[0].fillna("")
#     color_mapper = {
#         group: COLORS[i % len(COLORS)]
#         for i, group in enumerate(results["S. No Group"].unique())
#     }
#     prev_group = None

#     for i in range(len(results)):
#         curr_row = start
#         window_data = results.iloc[i]
#         qty_windows = value(window_data, "Area Qty (nos)")
#         col_values = {}

#         if product == "Louvers":
#             idx_mapper = LOUVER_PROFILE_COLUMNS
#             columns = ["Area Name", "S. No", "Orientation", "Height (mm)", "Width (mm)"]

#             col_values["No. of Pieces"] = value(
#                 window_data, "No. of Pieces", qty_windows
#             )
#             length_col = "Length (mm)"
#         elif product == "Beam C-Channel":
#             idx_mapper = BEAMC_PROFILE_COLUMNS
#             columns = ["Area Name", "S. No", "Width (mm)", "Length (mm)"]
#             length_col = "Beam Length (mm)"

#         col_values["Area Qty (nos)"] = qty_windows
#         for col_name in columns:
#             col_values[col_name] = value(window_data, col_name)

#         used = value(window_data, "Used Table")
#         waste = value(window_data, "Waste Table")

#         seen = set()
#         for piece_length in used:

#             if piece_length not in seen:
#                 piece_start_row = curr_row

#                 length = profile_xl.cell(row=curr_row, column=idx_mapper[length_col])
#                 set_cell(length, piece_length)
#                 seen.add(piece_length)

#                 qty = profile_xl.cell(
#                     row=curr_row, column=idx_mapper["Piece Qty (nos)"]
#                 )
#                 set_cell(qty, used[piece_length] * qty_windows)

#             if piece_length in waste:
#                 for i in range(len(waste[piece_length])):
#                     subarr = waste[piece_length][i]

#                     waste_length = profile_xl.cell(
#                         row=curr_row, column=idx_mapper["Scrap Length (mm)"]
#                     )
#                     set_cell(waste_length, subarr[0])

#                     waste_qty = profile_xl.cell(
#                         row=curr_row, column=idx_mapper["Scrap Qty (nos)"]
#                     )
#                     set_cell(waste_qty, subarr[1] * qty_windows)

#                     if i != 0:
#                         col_d = profile_xl.cell(
#                             row=curr_row, column=idx_mapper["Total Product Length (m)"]
#                         )
#                         col_d.value = None

#                     curr_row += 1

#                 curr_row -= 1

#                 # merge all rows with the same piece length
#                 for col in [length_col, "Piece Qty (nos)", "Total Product Length (m)"]:
#                     profile_xl.merge_cells(
#                         start_row=piece_start_row,
#                         start_column=idx_mapper[col],
#                         end_row=curr_row,
#                         end_column=idx_mapper[col],
#                     )

#             curr_row += 1
#         curr_row -= 1

#         if product == "Louvers":
#             merge_columns = [
#                 "Area Name",
#                 "S. No",
#                 "Orientation",
#                 "Height (mm)",
#                 "Width (mm)",
#                 "Area Qty (nos)",
#                 "Area (ft2)",
#                 "No. of Pieces",
#                 "Total Length (m)",
#             ]
#         elif product == "Beam C-Channel":
#             merge_columns = [
#                 "Area Name",
#                 "S. No",
#                 "Width (mm)",
#                 "Length (mm)",
#                 "Area Qty (nos)",
#                 "Total Length (m)",
#             ]

#         for col in merge_columns:
#             col_idx = idx_mapper[col]
#             profile_xl.merge_cells(
#                 start_row=start,
#                 start_column=col_idx,
#                 end_row=curr_row,
#                 end_column=col_idx,
#             )
#             if col in col_values:
#                 cell = profile_xl.cell(row=start, column=col_idx)
#                 set_cell(cell, col_values[col])

#         color = color_mapper[window_data["S. No Group"]]
#         fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
#         max_col_idx = 16 if product == "Louvers" else 13

#         for r in profile_xl.iter_rows(
#             min_row=start, max_row=curr_row, min_col=1, max_col=max_col_idx
#         ):
#             for cell in r:
#                 cell.fill = fill

#         if window_data["S. No Group"] != prev_group:
#             add_box_border(profile_xl, "medium", start, curr_row - 1, 1, max_col_idx)
#             prev_group = window_data["S. No Group"]

#         start = curr_row + 1

#     if product == "Louvers":
#         pitch = profile_xl.cell(row=1, column=2)
#         set_cell(pitch, value(window_data, "Pitch (mm)"))

#     project_title_col = 3 if product == "Louvers" else 1
#     project_title = profile_xl.cell(row=2, column=project_title_col)
#     set_cell(project_title, value(window_data, "Project Title"))

#     profile_xl = add_total_row(profile_xl, "profile", start, max_col_idx)
#     profile_xl.delete_rows(start + 1, profile_xl.max_row)

#     return profile_xl


# def generate_accessories_xl(product, start, carrier_xl, results):
#     output_xl, _, row_max, col_max = fill_window_data(
#         product, start, 2, carrier_xl, results
#     )
#     output_xl = add_total_row(output_xl, "carrier", row_max, col_max)
#     output_xl.delete_rows(row_max + 1, output_xl.max_row)
#     return output_xl


def generate_offer_xl(start, offer_xl, offer_df, offer_df_cols, common_vars):

    product = "Louvers"
    if offer_xl.cell(row=1, column=1).value == "Beam C-Channel":
        product = "Beam C-Channel"

    offer_xl, row_max, col_max = fill_offer_data(
        offer_xl, start, offer_df, offer_df_cols
    )

    zero_cols = [
        col
        for col, config in offer_df_cols.items()
        if config.get("hide_if_zero")
        and col in offer_df.columns
        and offer_df[col].sum() == 0
    ]

    if zero_cols:
        title_only = "Aerofoil Type" not in offer_df.columns
        offer_xl, col_max = remove_zero_total_cols(
            offer_xl,
            zero_cols,
            col_max,
            title_only,
            offer_df_cols
        )

    if product == "Louvers":
        pitch = offer_xl.cell(row=1, column=2)
        set_cell("offer", pitch, common_vars["pitch"])

    project_title_col = 3 if product == "Louvers" else 1
    project_title = offer_xl.cell(row=2, column=project_title_col)
    set_cell("offer", project_title, common_vars["project_title"])

    offer_xl = add_total_row(offer_xl, "offer", row_max, col_max)

    offer_xl.delete_rows(row_max + 1, offer_xl.max_row)

    return offer_xl


def generate_inventory_xl(inv_xl, inv_data):

    inv_xl, pvc_dv = add_dropdown(inv_xl, [item[1] for item in PVC_GITTY_TYPES])
    inv_xl, ft_dv = add_dropdown(inv_xl, [item[1] for item in FT_TYPES])
    inv_xl, sd_dv = add_dropdown(inv_xl, [item[1] for item in SELF_DRILLING_TYPES])
    inv_xl, bs_dv = add_dropdown(inv_xl, [item[1] for item in BLACK_GYPSUM])
    inv_xl, la_dv = add_dropdown(inv_xl, [item[1] for item in L_ANGLE])

    lookup = {
        "PVC GITTY": pvc_dv,
        "FULL THREADED SCREW": ft_dv,
        "SELF DRILLING": sd_dv,
        "BLACK GYPSUM DRYWALL SCREW": bs_dv,
        "L-ANGLE": la_dv,
    }
    idx_to_col = {0: 1, 1: 2, 2: 3, 3: 7, 4: 8, 5: 9, 6: 10, 7: 11, 8: 4}
    curr = 4

    for _, item in inv_data.iterrows():
        has_dropdown_in_row = False

        for col_idx, idx_value in enumerate(item):
            cell = inv_xl.cell(row=curr, column=idx_to_col[col_idx])

            if col_idx in [0, 1]:
                if col_idx == 1:  # Product Name column
                    # Check existing lookup dropdowns
                    for key, dropdown in lookup.items():
                        if key in idx_value:
                            dropdown.add(cell)
                            has_dropdown_in_row = True
                            break

                    # Product specific accessory dropdowns
                    if "COTTAL CORNER" in idx_value:
                        inv_xl, cottal_dv = add_dropdown(
                            inv_xl, [item[1] for item in COTTAL_CORNERS]
                        )
                        cottal_dv.add(cell)
                        lookup["COTTAL CORNER"] = cottal_dv
                        has_dropdown_in_row = True
                    elif "L-ANGLES" in idx_value:
                        inv_xl, langle_dv = add_dropdown(
                            inv_xl, [item[1] for item in L_ANGLE]
                        )
                        langle_dv.add(cell)
                        lookup["L-ANGLE"] = langle_dv
                        has_dropdown_in_row = True
                    elif "ENDCAP FIXTURE" in idx_value:
                        inv_xl, rect_ec_dv = add_dropdown(
                            inv_xl, [item[1] for item in RECT_EC_FIXING]
                        )
                        rect_ec_dv.add(cell)
                        lookup["ENDCAP FIXTURE"] = rect_ec_dv
                        has_dropdown_in_row = True
                    elif "FIXTURE" in idx_value:
                        inv_xl, fixing_dv = add_dropdown(
                            inv_xl, [item[1] for item in FIXING_TYPES]
                        )
                        fixing_dv.add(cell)
                        lookup["FIXTURE"] = fixing_dv
                        has_dropdown_in_row = True

                set_cell(
                    "inventory",
                    cell,
                    value=idx_value,
                    alignment="left" if col_idx == 1 else "center",
                )
            else:
                set_cell("inventory", cell, value=idx_value)

        # If this row has a dropdown, add VLOOKUP formula to Product Code column
        if has_dropdown_in_row:
            code_cell = inv_xl.cell(row=curr, column=1)  # Product Code column
            vlookup_formula = (
                f'=IFERROR(INDEX(ACCESSORIES!B:B,MATCH(B{curr},ACCESSORIES!A:A,0)),"")'
            )
            set_cell("inventory", code_cell, value=vlookup_formula, alignment="center")

        curr += 1

    # Add a row for a general accessory dropdown
    formula = "ACCESSORIES!$A:$A"
    accs_dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    inv_xl.add_data_validation(accs_dv)

    # Product Code cell with VLOOKUP formula for new accessory row
    code_cell = inv_xl.cell(row=curr, column=1)
    vlookup_formula = (
        f'=IFERROR(INDEX(ACCESSORIES!B:B,MATCH(B{curr},ACCESSORIES!A:A,0)),"")'
    )
    set_cell("inventory", code_cell, value=vlookup_formula, alignment="center")

    # Product Name cell with dropdown
    cell1 = inv_xl.cell(row=curr, column=2)
    set_cell("inventory", cell1, value="SELECT MORE ACCESSORIES", alignment="left")
    accs_dv.add(cell1)

    set_cell("inventory", inv_xl.cell(row=curr, column=7), value="pcs")

    tdate = inv_xl.cell(row=1, column=1)
    set_cell("inventory", tdate, value=date.today().strftime("%d-%m-%y"), size=14)

    curr += 1
    inv_xl.delete_rows(curr, inv_xl.max_row)

    return inv_xl


def generate_installer_xl(inst_xl, area_data, common_vars):

    def sc(cell_ref, value, bold=False, alignment="center", size=10):
        set_cell("installer", ws[cell_ref], value, bold=bold, alignment=alignment, size=size)

    tmp_dir = "/tmp/installer_pages"
    os.makedirs(tmp_dir, exist_ok=True)
    pdf_paths = []

    for idx, row in area_data.iterrows():
        wb = load_workbook(inst_xl)
        ws = wb.active

        s_no      = row.get("s_no", idx + 1)
        area_name = row.get("area_name", "")

        sc("A6",  common_vars.get("project_title", ""), bold=True)
        sc("A8",  f"Window {s_no} | {area_name}", bold=True)

        sc("E15", common_vars.get("product", ""))

        img = generate_window_image(row, common_vars)

        xl_img = XLImage(img)
        xl_img.width  = 570
        xl_img.height = 550
        xl_img.anchor = "A21"
        ws.add_image(xl_img)

        sc("A61",  common_vars.get("project_title", ""), bold=True)
        sc("A63",  f"Window {s_no} | {area_name}", bold=True)
        fill_cut_plan(ws, row["cut_plan"])

        xlsx_path = os.path.join(tmp_dir, f"window_{idx}.xlsx")
        # from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties

        # ws.sheet_properties = WorksheetProperties()
        # ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

        # ws.page_setup.paperSize  = ws.PAPERSIZE_A4
        # ws.page_setup.fitToWidth = 1
        # ws.page_setup.fitToHeight = 0
        # ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)
        # ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)
        wb.save(xlsx_path)

        lo_bin = get_libreoffice_path()
        # subprocess.run(
        #     [lo_bin, "--headless", "--convert-to", "pdf", "--outdir", tmp_dir, xlsx_path],
        #     check=True, capture_output=True
        # )
        subprocess.run(
            [lo_bin, "--headless", "--norestore",
            "--convert-to", "pdf",
            "--outdir", tmp_dir, xlsx_path],
            check=True, capture_output=True,
            env={**os.environ, "HOME": "/tmp"}
        )

        pdf_path = os.path.join(tmp_dir, f"window_{idx}.pdf")
        if os.path.exists(pdf_path):
            pdf_paths.append(pdf_path)

    writer = PdfWriter()
    for pdf_path in pdf_paths:
        for page in PdfReader(pdf_path).pages:
            writer.add_page(page)

    output_pdf = os.path.join(tmp_dir, "installer_all_windows.pdf")
    with open(output_pdf, "wb") as f:
        writer.write(f)

    inst_output = BytesIO()
    with open(output_pdf, "rb") as f:
        inst_output.write(f.read())
    inst_output.seek(0)
    return inst_output


def convert(product, results, common_vars):

    area_data = results[0]
    offer_xl_cols = results[1]
    offer_df = results[2]
    inventory_df = results[3]

    ext = ""
    if product in ["Aerofoil"]:
        if "Fixing Method" in results:
            col = "Fixing Method"
        else:
            col = "Installation Method"
        ext = "-" + results[col].iloc[0]

    template = get_xl_templates(product + ext, "profile")
    wb = openpyxl.load_workbook(template)

    offer_template = get_xl_templates(product + ext, "offer")
    offer_wb = openpyxl.load_workbook(offer_template, data_only=False)
    offer_xl = offer_wb.worksheets[0]

    start = 4
    offer_xl = generate_offer_xl(start, offer_xl, offer_df, offer_xl_cols, common_vars)

    # profile_xl = wb.worksheets[0]
    # start = 5
    # profile_xl = generate_profile_xl(start, profile_xl, results)

    # carrier_xl = wb.worksheets[1]
    # start = 3
    # carrier_xl = generate_accessories_xl(product, start, carrier_xl, results)

    inv_template = "reference_xls/inventory_xl.xlsx"
    inventory_wb = openpyxl.load_workbook(inv_template, data_only=False)
    inventory_xl = inventory_wb.worksheets[0]
    inventory_xl = generate_inventory_xl(inventory_xl, inventory_df)
    accs_xl = inventory_wb["ACCESSORIES"]
    accs_xl.sheet_state = "hidden"

    inst_template = "reference_xls/template.xlsx"
    # inst_wb = openpyxl.load_workbook(inst_template, data_only=False)
    # inst_xl = inst_wb.worksheets[0]
    inst_output = generate_installer_xl(inst_template, area_data, common_vars)

    # if product in ["Aerofoil", "S-Louvers", "Rectangular Louvers", "Cottal"]:
    #     ext = ""
    #     if product == "Aerofoil":
    #         ext = results["Aerofoil Type"].iloc[0]
    #     elif product in ["S-Louvers", "Rectangular Louvers", "Cottal"]:
    #         ext = results["Louver Size"].iloc[0]

    #     xls = {
    #         profile_xl: (f"Profile Calculation - {product} {ext}", 3),
    #         carrier_xl: (f"Accessories - {product} {ext}", 1),
    #         offer_xl: (f"{product} {ext}", 3),
    #         inventory_xl: (f"Inventory Sheet - {product} {ext}", 3),
    #     }

    #     for xl in xls:
    #         title = xl.cell(row=1, column=xls[xl][1])
    #         set_cell(title, xls[xl][0], bold=True, size=14)

    # profile_output = BytesIO()
    # wb.save(profile_output)
    # profile_output.seek(0)

    offer_output = BytesIO()
    offer_wb.save(offer_output)
    offer_output.seek(0)

    inv_output = BytesIO()
    inventory_wb.save(inv_output)
    inv_output.seek(0)

    return offer_output, inv_output, inst_output #profile_output, offer_output, inv_output
